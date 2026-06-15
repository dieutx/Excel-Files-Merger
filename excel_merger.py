from __future__ import annotations

from collections.abc import Callable, Iterable
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm"}
INVALID_SHEET_TITLE_CHARS = set("[]:*?/\\")
MAX_SHEET_TITLE_LENGTH = 31
DEDUPLICATED_TITLE_BASE_LENGTH = 20
MERGE_MODE_COMBINE = "combine"
MERGE_MODE_SEPARATE = "separate"
COMBINED_SHEET_NAME = "Combined"
SOURCE_FILE_COLUMN = "_source_file"
SOURCE_SHEET_COLUMN = "_source_sheet"


@dataclass(frozen=True)
class MergeResult:
    files_processed: int
    sheets_processed: int
    rows_written: int
    output_path: Path


@dataclass(frozen=True)
class MergeProgressEvent:
    event_type: str
    total_files: int
    current_file_index: int = 0
    file_path: Path | None = None
    sheet_title: str | None = None
    files_processed: int = 0
    sheets_processed: int = 0
    rows_written: int = 0
    output_path: Path | None = None


ProgressCallback = Callable[[MergeProgressEvent], None]


def discover_excel_files(folder: str | Path, exclude_path: str | Path | None = None) -> list[Path]:
    folder_path = Path(folder)
    excluded = Path(exclude_path).resolve() if exclude_path else None
    files: list[Path] = []

    for path in folder_path.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if excluded and path.resolve() == excluded:
            continue
        files.append(path)

    return sorted(files, key=lambda item: item.name.lower())


def sanitize_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    cleaned = "".join("_" if char in INVALID_SHEET_TITLE_CHARS else char for char in raw_title).strip()
    cleaned = cleaned.strip("'").strip()
    if not cleaned or set(cleaned) == {"_"}:
        cleaned = "Sheet"

    base = cleaned[:MAX_SHEET_TITLE_LENGTH]
    candidate = base
    counter = 2
    used_title_keys = {title.casefold() for title in used_titles}

    while candidate.casefold() in used_title_keys:
        suffix = f" ({counter})"
        base_length = min(DEDUPLICATED_TITLE_BASE_LENGTH, MAX_SHEET_TITLE_LENGTH - len(suffix))
        candidate = f"{base[:base_length]}{suffix}"
        counter += 1

    used_titles.add(candidate)
    return candidate


def merge_excel_files(
    folder: str | Path,
    output_path: str | Path,
    mode: str,
    on_progress: ProgressCallback | None = None,
) -> MergeResult:
    output = Path(output_path)
    files = discover_excel_files(folder, exclude_path=output)
    if not files:
        raise ValueError("No Excel files were found in the selected folder.")

    if on_progress:
        on_progress(MergeProgressEvent("scan_complete", total_files=len(files)))

    if mode == MERGE_MODE_COMBINE:
        return _merge_combined(files, output, on_progress)
    if mode == MERGE_MODE_SEPARATE:
        return _merge_separate(files, output, on_progress)
    raise ValueError(f"Unsupported merge mode: {mode}")


def _merge_combined(
    files: Iterable[Path],
    output: Path,
    on_progress: ProgressCallback | None = None,
) -> MergeResult:
    file_list = list(files)
    total_files = len(file_list)
    records: list[dict[str, object]] = []
    data_columns: list[str] = []
    files_processed = 0
    sheets_processed = 0

    for file_index, file_path in enumerate(file_list, start=1):
        if on_progress:
            on_progress(
                MergeProgressEvent(
                    "file_start",
                    total_files=total_files,
                    current_file_index=file_index,
                    file_path=file_path,
                    files_processed=files_processed,
                    sheets_processed=sheets_processed,
                    rows_written=len(records),
                )
            )
        files_processed += 1
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                rows = _non_empty_rows(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                sheets_processed += 1
                headers = _normalize_headers(rows[0], reserved_names={SOURCE_FILE_COLUMN, SOURCE_SHEET_COLUMN})
                for header in headers:
                    if header not in data_columns:
                        data_columns.append(header)

                for row in rows[1:]:
                    record = {
                        SOURCE_FILE_COLUMN: file_path.name,
                        SOURCE_SHEET_COLUMN: sheet.title,
                    }
                    for header, value in zip(headers, row):
                        record[header] = value
                    records.append(record)

                if on_progress:
                    on_progress(
                        MergeProgressEvent(
                            "sheet_processed",
                            total_files=total_files,
                            current_file_index=file_index,
                            file_path=file_path,
                            sheet_title=sheet.title,
                            files_processed=files_processed,
                            sheets_processed=sheets_processed,
                            rows_written=len(records),
                        )
                    )
        finally:
            workbook.close()

        if on_progress:
            on_progress(
                MergeProgressEvent(
                    "file_complete",
                    total_files=total_files,
                    current_file_index=file_index,
                    file_path=file_path,
                    files_processed=files_processed,
                    sheets_processed=sheets_processed,
                    rows_written=len(records),
                )
            )

    if not records:
        raise ValueError("The selected Excel files do not contain any data rows.")

    output.parent.mkdir(parents=True, exist_ok=True)
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = COMBINED_SHEET_NAME

    columns = [SOURCE_FILE_COLUMN, SOURCE_SHEET_COLUMN, *data_columns]
    output_sheet.append(columns)
    for record in records:
        output_sheet.append([record.get(column) for column in columns])

    if on_progress:
        on_progress(
            MergeProgressEvent(
                "writing_output",
                total_files=total_files,
                current_file_index=total_files,
                files_processed=files_processed,
                sheets_processed=sheets_processed,
                rows_written=len(records),
                output_path=output,
            )
        )

    output_workbook.save(output)
    output_workbook.close()
    result = MergeResult(files_processed, sheets_processed, len(records), output)
    if on_progress:
        on_progress(
            MergeProgressEvent(
                "complete",
                total_files=total_files,
                current_file_index=total_files,
                files_processed=result.files_processed,
                sheets_processed=result.sheets_processed,
                rows_written=result.rows_written,
                output_path=result.output_path,
            )
        )
    return result


def _merge_separate(
    files: Iterable[Path],
    output: Path,
    on_progress: ProgressCallback | None = None,
) -> MergeResult:
    file_list = list(files)
    total_files = len(file_list)
    output_workbook = Workbook()
    default_sheet = output_workbook.active
    output_workbook.remove(default_sheet)

    used_titles: set[str] = set()
    files_processed = 0
    sheets_processed = 0
    rows_written = 0

    try:
        for file_index, file_path in enumerate(file_list, start=1):
            if on_progress:
                on_progress(
                    MergeProgressEvent(
                        "file_start",
                        total_files=total_files,
                        current_file_index=file_index,
                        file_path=file_path,
                        files_processed=files_processed,
                        sheets_processed=sheets_processed,
                        rows_written=rows_written,
                    )
                )
            files_processed += 1
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                for sheet in workbook.worksheets:
                    rows = _non_empty_rows(sheet.iter_rows(values_only=True))
                    if not rows:
                        continue

                    title = sanitize_sheet_title(f"{file_path.stem} - {sheet.title}", used_titles)
                    output_sheet = output_workbook.create_sheet(title)
                    for row in rows:
                        output_sheet.append(row)

                    sheets_processed += 1
                    rows_written += len(rows)

                    if on_progress:
                        on_progress(
                            MergeProgressEvent(
                                "sheet_processed",
                                total_files=total_files,
                                current_file_index=file_index,
                                file_path=file_path,
                                sheet_title=sheet.title,
                                files_processed=files_processed,
                                sheets_processed=sheets_processed,
                                rows_written=rows_written,
                            )
                        )
            finally:
                workbook.close()

            if on_progress:
                on_progress(
                    MergeProgressEvent(
                        "file_complete",
                        total_files=total_files,
                        current_file_index=file_index,
                        file_path=file_path,
                        files_processed=files_processed,
                        sheets_processed=sheets_processed,
                        rows_written=rows_written,
                    )
                )

        if sheets_processed == 0:
            raise ValueError("The selected Excel files do not contain any usable sheets.")

        output.parent.mkdir(parents=True, exist_ok=True)
        if on_progress:
            on_progress(
                MergeProgressEvent(
                    "writing_output",
                    total_files=total_files,
                    current_file_index=total_files,
                    files_processed=files_processed,
                    sheets_processed=sheets_processed,
                    rows_written=rows_written,
                    output_path=output,
                )
            )

        output_workbook.save(output)
        result = MergeResult(files_processed, sheets_processed, rows_written, output)
        if on_progress:
            on_progress(
                MergeProgressEvent(
                    "complete",
                    total_files=total_files,
                    current_file_index=total_files,
                    files_processed=result.files_processed,
                    sheets_processed=result.sheets_processed,
                    rows_written=result.rows_written,
                    output_path=result.output_path,
                )
            )
        return result
    finally:
        output_workbook.close()


def _non_empty_rows(rows: Iterable[tuple[object, ...]]) -> list[tuple[object, ...]]:
    return [tuple(row) for row in rows if any(value is not None for value in row)]


def _normalize_headers(row: tuple[object, ...], reserved_names: set[str] | None = None) -> list[str]:
    reserved_names = reserved_names or set()
    base_headers: list[tuple[str, bool]] = []
    real_header_bases: set[str] = set()

    for index, value in enumerate(row, start=1):
        text = str(value).strip() if value is not None else ""
        is_blank = not text
        base = text if text else f"Column {index}"
        base_headers.append((base, is_blank))
        if not is_blank:
            real_header_bases.add(base)

    reserved_bases = {base for base, _ in base_headers}
    headers: list[str] = []
    used: set[str] = set()

    for base, is_blank in base_headers:
        if base not in used and base not in reserved_names and not (is_blank and base in real_header_bases):
            header = base
        else:
            counter = 2
            header = f"{base} {counter}"
            while header in used or header in reserved_bases or header in reserved_names:
                counter += 1
                header = f"{base} {counter}"
        used.add(header)
        headers.append(header)

    return headers
