from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_merger import (
    MERGE_MODE_COMBINE,
    MERGE_MODE_SEPARATE,
    discover_excel_files,
    merge_excel_files,
    sanitize_sheet_title,
)


def make_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for title, rows in sheets.items():
        sheet = workbook.create_sheet(title)
        for row in rows:
            sheet.append(row)

    workbook.save(path)


def read_sheet_rows(path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook[sheet_name]
    return [tuple(row) for row in sheet.iter_rows(values_only=True)]


def test_discovers_supported_excel_files_and_ignores_temp_files(tmp_path: Path):
    (tmp_path / "b.xlsx").write_bytes(b"")
    (tmp_path / "a.xlsm").write_bytes(b"")
    (tmp_path / "~$locked.xlsx").write_bytes(b"")
    (tmp_path / "notes.csv").write_text("not excel", encoding="utf-8")
    (tmp_path / "nested").mkdir()

    files = discover_excel_files(tmp_path)

    assert [path.name for path in files] == ["a.xlsm", "b.xlsx"]


def test_discovery_can_exclude_output_file_inside_source_folder(tmp_path: Path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "combined.xlsx"
    source.write_bytes(b"")
    output.write_bytes(b"")

    files = discover_excel_files(tmp_path, exclude_path=output)

    assert [path.name for path in files] == ["source.xlsx"]


def test_sanitizes_and_deduplicates_excel_sheet_titles():
    used: set[str] = set()

    first = sanitize_sheet_title("2026/Q1:Report*East[Raw]", used)
    second = sanitize_sheet_title("2026/Q1:Report*East[Raw]", used)
    blank = sanitize_sheet_title("[]:*?/\\", used)

    assert first == "2026_Q1_Report_East_Raw_"
    assert second == "2026_Q1_Report_East_ (2)"
    assert blank == "Sheet"
    assert all(len(title) <= 31 for title in used)


def test_sanitizes_sheet_titles_deduplicates_case_insensitively():
    used: set[str] = set()

    first = sanitize_sheet_title("Report", used)
    second = sanitize_sheet_title("report", used)

    assert first == "Report"
    assert second == "report (2)"


def test_sanitizes_sheet_titles_by_stripping_outer_apostrophes_and_whitespace():
    used: set[str] = set()

    assert sanitize_sheet_title("'  Report  '", used) == "Report"
    assert sanitize_sheet_title("  'Quarterly Report'  ", used) == "Quarterly Report"


def test_blank_and_whitespace_sheet_titles_default_to_sheet():
    assert sanitize_sheet_title("", set()) == "Sheet"
    assert sanitize_sheet_title("   ", set()) == "Sheet"
    assert sanitize_sheet_title("'   '", set()) == "Sheet"


def test_sanitizes_long_sheet_titles_to_excel_limit_and_deduplicates():
    used: set[str] = set()
    raw_title = "ABCDEFGHIJKLMNOPQRSTUVWXYZ1234567890"

    first = sanitize_sheet_title(raw_title, used)
    second = sanitize_sheet_title(raw_title, used)

    assert len(raw_title) > 31
    assert first == "ABCDEFGHIJKLMNOPQRSTUVWXYZ12345"
    assert len(first) == 31
    assert second == "ABCDEFGHIJKLMNOPQRST (2)"
    assert all(len(title) <= 31 for title in used)


def test_combines_workbooks_into_one_sheet_with_source_metadata(tmp_path: Path):
    make_workbook(
        tmp_path / "a_sales.xlsx",
        {"North": [["Name", "Amount"], ["Ada", 10], ["Ben", 20]]},
    )
    make_workbook(
        tmp_path / "b_returns.xlsx",
        {"South": [["Name", "Reason"], ["Cora", "Damaged"]]},
    )
    output = tmp_path / "combined.xlsx"

    result = merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)

    rows = read_sheet_rows(output, "Combined")
    assert result.files_processed == 2
    assert result.sheets_processed == 2
    assert result.rows_written == 3
    assert rows == [
        ("_source_file", "_source_sheet", "Name", "Amount", "Reason"),
        ("a_sales.xlsx", "North", "Ada", 10, None),
        ("a_sales.xlsx", "North", "Ben", 20, None),
        ("b_returns.xlsx", "South", "Cora", None, "Damaged"),
    ]


def test_combined_mode_emits_progress_events(tmp_path: Path):
    make_workbook(
        tmp_path / "a_sales.xlsx",
        {"North": [["Name"], ["Ada"]]},
    )
    make_workbook(
        tmp_path / "b_sales.xlsx",
        {"South": [["Name"], ["Ben"]]},
    )
    output = tmp_path / "combined.xlsx"
    events = []

    result = merge_excel_files(
        tmp_path,
        output,
        MERGE_MODE_COMBINE,
        on_progress=events.append,
    )

    event_types = [event.event_type for event in events]
    assert event_types == [
        "scan_complete",
        "file_start",
        "sheet_processed",
        "file_complete",
        "file_start",
        "sheet_processed",
        "file_complete",
        "writing_output",
        "complete",
    ]
    assert [event.current_file_index for event in events if event.event_type == "file_start"] == [1, 2]
    assert all(event.total_files == 2 for event in events)
    assert events[-1].files_processed == result.files_processed == 2
    assert events[-1].sheets_processed == result.sheets_processed == 2
    assert events[-1].rows_written == result.rows_written == 2
    assert events[-1].output_path == output


def test_combined_mode_writes_output_when_progress_callback_raises(tmp_path: Path):
    make_workbook(
        tmp_path / "a_sales.xlsx",
        {"North": [["Name"], ["Ada"]]},
    )
    output = tmp_path / "combined.xlsx"

    def raise_on_writing_output(event):
        if event.event_type == "writing_output":
            raise RuntimeError("progress failed")

    result = merge_excel_files(
        tmp_path,
        output,
        MERGE_MODE_COMBINE,
        on_progress=raise_on_writing_output,
    )

    rows = read_sheet_rows(output, "Combined")
    assert output.exists()
    assert result.rows_written == 1
    assert rows == [
        ("_source_file", "_source_sheet", "Name"),
        ("a_sales.xlsx", "North", "Ada"),
    ]


def test_keeps_source_sheets_separate_with_safe_unique_names(tmp_path: Path):
    make_workbook(
        tmp_path / "a_east.xlsx",
        {"Raw Data": [["Name"], ["Ada"]]},
    )
    make_workbook(
        tmp_path / "b_east_copy.xlsx",
        {"Raw Data": [["Name"], ["Ben"]]},
    )
    output = tmp_path / "separate.xlsx"

    result = merge_excel_files(tmp_path, output, MERGE_MODE_SEPARATE)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == ["a_east - Raw Data", "b_east_copy - Raw Data"]
    assert read_sheet_rows(output, "a_east - Raw Data") == [("Name",), ("Ada",)]
    assert read_sheet_rows(output, "b_east_copy - Raw Data") == [("Name",), ("Ben",)]
    assert result.files_processed == 2
    assert result.sheets_processed == 2
    assert result.rows_written == 4


def test_separate_mode_emits_progress_events(tmp_path: Path):
    make_workbook(
        tmp_path / "a_east.xlsx",
        {"Raw": [["Name"], ["Ada"]]},
    )
    make_workbook(
        tmp_path / "b_west.xlsx",
        {"Raw": [["Name"], ["Ben"]]},
    )
    output = tmp_path / "separate.xlsx"
    events = []

    result = merge_excel_files(
        tmp_path,
        output,
        MERGE_MODE_SEPARATE,
        on_progress=events.append,
    )

    event_types = [event.event_type for event in events]
    assert event_types == [
        "scan_complete",
        "file_start",
        "sheet_processed",
        "file_complete",
        "file_start",
        "sheet_processed",
        "file_complete",
        "writing_output",
        "complete",
    ]
    assert [event.current_file_index for event in events if event.event_type == "file_complete"] == [1, 2]
    assert all(event.total_files == 2 for event in events)
    assert events[-1].files_processed == result.files_processed == 2
    assert events[-1].sheets_processed == result.sheets_processed == 2
    assert events[-1].rows_written == result.rows_written == 4
    assert events[-1].output_path == output


def test_separate_mode_fails_cleanly_when_no_usable_sheets_exist(tmp_path: Path):
    make_workbook(tmp_path / "empty.xlsx", {"Empty": []})
    output = tmp_path / "separate.xlsx"

    try:
        merge_excel_files(tmp_path, output, MERGE_MODE_SEPARATE)
    except ValueError as exc:
        assert str(exc) == "The selected Excel files do not contain any usable sheets."
    else:
        raise AssertionError("Expected ValueError when no usable sheets exist")


def test_combined_mode_preserves_headers_that_collide_with_generated_duplicates(tmp_path: Path):
    make_workbook(
        tmp_path / "headers.xlsx",
        {"Data": [["Name", "Name 2", "Name"], ["first", "second", "third"]]},
    )
    output = tmp_path / "combined.xlsx"

    merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)

    rows = read_sheet_rows(output, "Combined")
    assert rows == [
        ("_source_file", "_source_sheet", "Name", "Name 2", "Name 3"),
        ("headers.xlsx", "Data", "first", "second", "third"),
    ]


def test_combined_mode_preserves_source_metadata_when_data_headers_collide(tmp_path: Path):
    make_workbook(
        tmp_path / "metadata_headers.xlsx",
        {"Data": [["_source_file", "_source_sheet", "Name"], ["file value", "sheet value", "Ada"]]},
    )
    output = tmp_path / "combined.xlsx"

    merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)

    rows = read_sheet_rows(output, "Combined")
    assert rows == [
        ("_source_file", "_source_sheet", "_source_file 2", "_source_sheet 2", "Name"),
        ("metadata_headers.xlsx", "Data", "file value", "sheet value", "Ada"),
    ]


def test_combined_mode_preserves_later_real_headers_that_match_generated_duplicates(tmp_path: Path):
    make_workbook(
        tmp_path / "headers.xlsx",
        {"Data": [["Name", "Name", "Name 2"], ["first", "second", "third"]]},
    )
    output = tmp_path / "combined.xlsx"

    merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)

    rows = read_sheet_rows(output, "Combined")
    assert rows == [
        ("_source_file", "_source_sheet", "Name", "Name 3", "Name 2"),
        ("headers.xlsx", "Data", "first", "second", "third"),
    ]


def test_combined_mode_preserves_later_real_headers_that_match_blank_generated_names(tmp_path: Path):
    make_workbook(
        tmp_path / "headers.xlsx",
        {"Data": [[None, "Column 1", None], ["first", "second", "third"]]},
    )
    output = tmp_path / "combined.xlsx"

    merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)

    rows = read_sheet_rows(output, "Combined")
    assert rows == [
        ("_source_file", "_source_sheet", "Column 1 2", "Column 1", "Column 3"),
        ("headers.xlsx", "Data", "first", "second", "third"),
    ]


def test_combined_mode_fails_cleanly_when_no_excel_files_exist(tmp_path: Path):
    output = tmp_path / "combined.xlsx"

    try:
        merge_excel_files(tmp_path, output, MERGE_MODE_COMBINE)
    except ValueError as exc:
        assert str(exc) == "No Excel files were found in the selected folder."
    else:
        raise AssertionError("Expected ValueError for an empty source folder")


def test_rejects_unknown_merge_mode(tmp_path: Path):
    make_workbook(tmp_path / "source.xlsx", {"Sheet1": [["Name"], ["Ada"]]})

    try:
        merge_excel_files(tmp_path, tmp_path / "out.xlsx", "unknown")
    except ValueError as exc:
        assert str(exc) == "Unsupported merge mode: unknown"
    else:
        raise AssertionError("Expected ValueError for unsupported mode")
